import openpyxl

def writeExcel(file_name, file_name_database):
        import_wb = openpyxl.load_workbook(file_name)
        import_sheet = import_wb['Sheet1']

        wb = openpyxl.load_workbook(file_name_database)
        sheet = wb['Первый лист']

        for i in range(2, import_sheet.max_row + 1):
            import_list_names = [import_sheet.cell(row=i, column=j).value for j in range(1, import_sheet.max_column + 1) if import_sheet.cell(row=i, column=j).value]

            for i in range(2, sheet.max_row + 1):
                list_names = [sheet.cell(row=i, column=j).value for j in range(1, sheet.max_column + 1) if sheet.cell(row=i, column=j).value]

                if(import_list_names == list_names): break
                elif(i == sheet.max_row):
                    max_row_value = sheet.max_row

                    for j in range(len(import_list_names)):
                        sheet.cell(row=max_row_value + 1, column=j + 1, value=import_list_names[j])

        wb.save(file_name_database)
        wb.close()
        import_wb.close()

def excelFormat(file_name_export, file_name_database):
    wb = openpyxl.load_workbook(file_name_database)
    sheet = wb['Первый лист']

    export_wb = openpyxl.Workbook()
    export_sheet = export_wb.active

    for i in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=i, column=j)
            export_sheet.cell(row=i, column=j, value=cell.value)

    export_wb.save(f"{file_name_export}.xlsx")
    export_wb.close()
    wb.close()