import openpyxl
import os 
import sqlite3
from converter_text import writeText, textFormat
from converter_excel import writeExcel, excelFormat
from converter_SQLite import writeSQL, sqlFormat
    
"""Функция выполняющая поиск совпадений по полученному слову. Возвращает список."""
def searchContact(word):
    wb = openpyxl.load_workbook('Work_7\excel.xlsx')
    sheet = wb['Первый лист']
    
    result = []

    for i in range(2, sheet.max_row + 1):
        list_name = [sheet.cell(row=i, column=j).value for j in range(1, sheet.max_column + 1) if sheet.cell(row=i, column=j).value]
        if(word in list_name): result.append(list_name)

    wb.close()
    return result

"""Функция добавления нового контакта в базу данных. На выходе получаем обновленный файл .xlsx."""
def addContact(parametrs):
    wb = openpyxl.load_workbook('Work_7\excel.xlsx')
    sheet = wb['Первый лист']
    end_row = sheet.max_row + 1

    for i in range(len(parametrs)):
        sheet.cell(row=end_row, column=i + 1, value=parametrs[i])
    
    wb.save('Work_7\excel.xlsx')
    wb.close()
    
"""Функция удаления контакта из базы данных. На выходе получаем обновленный файл .xlsx. На вход принимает имя контакта."""
def deleteContact(contact):
    wb = openpyxl.load_workbook('Work_7\excel.xlsx')
    sheet = wb['Первый лист']

    for i in range(2, sheet.max_row + 1):
        for i in range(2, sheet.max_row + 1):
            list_names = [sheet.cell(row=i, column=j).value for j in range(1, sheet.max_column + 1) if sheet.cell(row=i, column=j).value]
        result = ','.join(list_names).replace(',', ' ') + "\n"

        if(result == contact): 
            sheet.delete_rows(i)
            break
        
    wb.save('Work_7\excel.xlsx')
    wb.close()

"""Функция подгрузки базы контактов из файла .xlsx. На вход принимает номер строки, которую считываем."""
def readDatabase(row):
    if (os.path.exists("Work_7\excel.xlsx")):
        wb = openpyxl.load_workbook('Work_7\excel.xlsx')
        sheet = wb['Первый лист']

        list_names = [sheet.cell(row=row, column=i).value for i in range(1, sheet.max_column + 1) if sheet.cell(row=row, column=i).value]

        return map(str, list_names)
                
"""Функция возвращает количество строк в базе, либо создает файл заново."""
def sizeDatabase():
    if (os.path.exists("Work_7\excel.xlsx")):
        wb = openpyxl.load_workbook('Work_7\excel.xlsx')
        sheet = wb['Первый лист']

    else:
        wb = openpyxl.Workbook()
        wb.create_sheet(title = 'Первый лист', index = 0)
        sheet = wb['Первый лист']
        parametrs = ["Фамилия", "Имя", "Отчество", "Рабочий телефон", "Дополнительный телефон"]

        for i in range(len(parametrs)):
            sheet.cell(row=1, column=i + 1, value=parametrs[i])
            
        wb.save('Work_7\excel.xlsx')

    row = int(sheet.max_row)   
    wb.close()

    return row

"""Функция сортировки базы. Построчно."""
def sortedData():
    wb = openpyxl.load_workbook('Work_7\excel.xlsx')
    sheet = wb['Первый лист']

    sort_wb = openpyxl.Workbook()
    sort_wb.create_sheet(title = 'Первый лист', index = 0)
    sort_sheet = sort_wb['Первый лист']
    parametrs = ["Фамилия", "Имя", "Отчество", "Рабочий телефон", "Дополнительный телефон"]
    

    for i in range(len(parametrs)):
        sort_sheet.cell(row=1, column=i + 1, value=parametrs[i])
        
    wb.save('Work_7\excel.xlsx')
    wb.close()

    
    count = sheet.max_row
    while (count >= 2):
        size_list = sort_sheet.max_row
        min_cell = sheet.cell(row=2, column=1).value
        min_list = []
        min_row = 2
        
        for i in range(2, sheet.max_row + 1):
            list_names = [sheet.cell(row=i, column=j).value for j in range(1, sheet.max_column + 1) if sheet.cell(row=i, column=j).value]
            surname = list_names[0]

            if(surname <= min_cell):
                min_cell = sheet.cell(row=i, column=1).value
                min_list = list_names
                min_row = i

        sheet.delete_rows(min_row)
        count -= 1

        for i in range(len(min_list)):
            sort_sheet.cell(row=size_list + 1, column=i+1, value=min_list[i])

        sort_wb.save("Work_7\excel.xlsx")

"""Функция импорта данных из файлов txt, xlsx, bd."""
def importData(file_name, file_name_database):
    match os.path.splitext(file_name)[1]:
        case ".xlsx": 
            writeExcel(file_name, file_name_database)
        case ".txt":
            writeText(file_name, file_name_database)
        case ".db": 
            writeSQL(file_name, file_name_database)

"""Функция експорта данных в файлы txt, xlsx, bd"""
def exportData(format, file_name_export, file_name_database):
    match format:
        case "txt": textFormat(file_name_export, file_name_database)
        case "xlsx": excelFormat(file_name_export, file_name_database)
        case "db": sqlFormat(file_name_export, file_name_database)