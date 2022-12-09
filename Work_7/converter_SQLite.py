import openpyxl
import sqlite3

def writeSQL(file_name):
    con = sqlite3.connect(file_name)
    cur = con.cursor()

    cur.execute("SELECT * FROM Sheet1")
    
    while True:
        contact = cur.fetchmany()
        if(not contact): break

        contact = [parametr for parametr in list(contact[0]) if parametr]

        wb = openpyxl.load_workbook('excel.xlsx')
        sheet = wb['Первый лист']

        for i in range(2, sheet.max_row + 1):
            list_names = [sheet.cell(row=i, column=j).value for j in range(1, sheet.max_column + 1) if sheet.cell(row=i, column=j).value]

            if(contact == list_names): break
            elif(i == sheet.max_row):
                max_row_value = sheet.max_row

                for j in range(len(contact)):
                    sheet.cell(row=max_row_value + 1, column=j + 1, value=contact[j])

        wb.save('excel.xlsx')
        wb.close()
    
    con.close()

def sqlFormat():
        wb = openpyxl.load_workbook('excel.xlsx')
        sheet = wb['Первый лист']
        
        con = sqlite3.connect('result.db')
        cur = con.cursor()
        
        cur.execute("""CREATE TABLE IF NOT EXISTS Sheet1('Фамилия' TXT,  'Имя' TEXT, 'Отчество' TEXT, 'Рабочий_телефон' TEXT, 'Дополнительный_телефон' TEXT);""")
        con.commit()

        for i in range(2, sheet.max_row + 1):
            list_names = [str(sheet.cell(row=i, column=j).value) for j in range(1, sheet.max_column + 1)]
            cur.execute("INSERT INTO Sheet1 VALUES(?, ?, ?, ?, ?);", list_names)
            con.commit()

        wb.close()