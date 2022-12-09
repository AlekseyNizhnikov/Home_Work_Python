import openpyxl

def writeText(file_name, file_name_database):
        with open(file_name, "r", encoding="UTF-8") as file:  
            contact = file.readline()
            
            while True:
                contact = file.readline()
                if (not contact): break
                contact = contact.split()

                wb = openpyxl.load_workbook(file_name_database)
                sheet = wb['Первый лист']
                
                if(sheet.max_row > 2):
                    for i in range(2, sheet.max_row + 1):
                        list_names = [sheet.cell(row=i, column=j).value for j in range(1, sheet.max_column + 1) if sheet.cell(row=i, column=j).value]

                        if(contact == list_names): break
                        elif(i == sheet.max_row):
                            max_row_value = sheet.max_row
                            for j in range(len(contact)):
                                sheet.cell(row=max_row_value + 1, column=j + 1, value=contact[j])
                else:
                    max_row_value = sheet.max_row
                    for j in range(len(contact)):
                        sheet.cell(row=max_row_value + 1, column=j + 1, value=contact[j])
                        
                wb.save(file_name_database)
                wb.close()

def textFormat(file_name_export, file_name_database):
        with open(f"{file_name_export}.txt", "w", encoding="UTF-8") as file:
            wb = openpyxl.load_workbook(file_name_database)
            sheet = wb['Первый лист']

            for i in range(2, sheet.max_row + 1):
                list_names = [sheet.cell(row=i, column=j).value for j in range(1, sheet.max_column + 1) if sheet.cell(row=i, column=j).value]
                
                file.write(f"{','.join(map(str, list_names)).replace(',', ' ')}\n")

            wb.close()